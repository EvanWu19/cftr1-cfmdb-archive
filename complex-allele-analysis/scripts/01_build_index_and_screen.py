"""
Step 1 of the CFTR1 complex-allele screen.

  (a) Build the partner-resolution alias index from CFTR1 (data/mutations.json,
      already in this repo) + the CFTR2 variant list (external file, see below).
  (b) Derive a Nucleotide-change and Consequence field for every variant.
  (c) Keyword-screen all 2,121 variants' Other Details + Submitted Phenotype
      Details for complex-allele language, producing the candidate set.
  (d) Write per-batch JSON files for the AI judgement pass (step 2 workflows).

Outputs (into ../intermediate/ and ../intermediate/batches*/):
  alias_index.json, all_records.json, candidates.json, notext_ids.json,
  batches/*.json (keyword candidates), batches2/*.json (non-keyword w/ text).

External input:
  CFTR2 variant spreadsheet ("CFTR2_<date>.xlsx"). NOT redistributed here; obtain
  the current file from https://cftr2.org (Downloads -> variant list). Pass its
  path with --cftr2 or set CFTR2_XLSX. The index embeds CFTR2 cross-references
  (legacy/protein/cDNA/alt-name) for partner resolution only.

Usage:
  python 01_build_index_and_screen.py --cftr2 /path/to/CFTR2_30January2026.xlsx
"""
import json, re, os, sys, argparse, shutil

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
INTER = os.path.join(HERE, "..", "intermediate")
MUTATIONS = os.path.join(REPO, "data", "mutations.json")

aa3to1 = {'Ala':'A','Arg':'R','Asn':'N','Asp':'D','Cys':'C','Gln':'Q','Glu':'E','Gly':'G',
          'His':'H','Ile':'I','Leu':'L','Lys':'K','Met':'M','Phe':'F','Pro':'P','Ser':'S',
          'Thr':'T','Trp':'W','Tyr':'Y','Val':'V','Ter':'X','Sec':'U'}


def norm_key(s):
    if not s: return None
    s = str(s).strip().replace('->', '>').replace('/', '>')
    s = re.sub(r'\s+', '', s)
    s = s.replace('[delta]', '').replace('delta', '').replace('Δ', '').replace('(', '').replace(')', '')
    return s.lower()


def prot_variants(p):
    out = set()
    if not p: return out
    out.add(norm_key(p))
    m = re.match(r'p?\.?\(?([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2}|=|X|\*|fs.*|Ter)', p.strip())
    if m and m.group(1) in aa3to1:
        a1 = aa3to1[m.group(1)]; num = m.group(2); rest = m.group(3)
        a2 = aa3to1.get(rest, rest)
        out.add(norm_key(a1 + num + a2))
    return out


def build_index(cftr1, cftr2_xlsx):
    index = {}

    def add(alias, cdna):
        k = norm_key(alias)
        if k and cdna:
            index.setdefault(k, set()).add(cdna)

    for m in cftr1:
        c = m.get('cdna_name')
        if not c: continue
        add(c, c)
        add(m.get('legacy_name'), c)
        for pv in prot_variants(m.get('protein_name')):
            if pv: index.setdefault(pv, set()).add(c)

    cftr2 = []
    if cftr2_xlsx and os.path.exists(cftr2_xlsx):
        import openpyxl
        wb = openpyxl.load_workbook(cftr2_xlsx, read_only=True)
        ws = wb['CFTR2 variants by legacy name']
        for r in ws.iter_rows(min_row=13, values_only=True):
            legacy, protein, cdna, alt = r[0], r[1], r[2], r[3]
            det = r[7]
            if not cdna: continue
            cftr2.append({'legacy': legacy, 'protein': protein, 'cdna': cdna, 'alt': alt, 'determination': det})
            add(cdna, cdna); add(legacy, cdna)
            for pv in prot_variants(protein):
                if pv: index.setdefault(pv, set()).add(cdna)
            if alt:
                for a in re.split(r'[;,]', str(alt)):
                    add(a.strip(), cdna)
        ws2 = wb['Genomic coordinates']
        for r in ws2.iter_rows(min_row=2, values_only=True):
            cdna, legacy = r[0], r[1]
            if cdna: add(cdna, cdna); add(legacy, cdna)
    else:
        sys.stderr.write("WARNING: no CFTR2 xlsx given; index built from CFTR1 only "
                         "(partner resolution will be weaker).\n")
    return {k: sorted(v) for k, v in index.items()}, cftr2


def derive_consequence(m):
    c = (m.get('cdna_name') or ''); p = (m.get('protein_name') or ''); region = (m.get('region') or '')
    if 'fs' in p or 'fsX' in p or 'fsTer' in p: return 'frameshift'
    if re.search(r'X\d*$', p) or 'Ter' in p or p.endswith('*'):
        return 'nonsense' if '>' in c else 'nonsense/truncating'
    if p.endswith('=') or 'p.(=)' in p or p == 'p.=': return 'synonymous'
    mm = re.match(r'p\.([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})', p)
    if mm and mm.group(1) != mm.group(3): return 'missense'
    if 'del' in c and 'ins' in c: return 'indel'
    if 'del' in c: return 'deletion (in-frame)' if ('del' in p and 'fs' not in p) else 'deletion'
    if 'ins' in c: return 'insertion'
    if 'dup' in c: return 'duplication'
    if re.search(r'c\.[-\d]+[-+]\d+', c) or region.startswith('intron'): return 'splicing (putative)'
    if '>' in c and not p: return 'substitution (non-coding/unknown)'
    if '>' in c: return 'substitution'
    return 'other/unspecified'


KEYWORD_PATTERNS = [
    ('complex allele', r'complex allele'), ('complex', r'\bcomplex\b'),
    ('same allele', r'same allele'), ('in cis', r'\bin cis\b'), ('cis', r'\bcis\b'),
    ('together', r'\btogether\b'), ('same chromosome', r'same chromosome'),
    ('same (non-)F508 chromosome', r'same non-?\[?delta\]?\s*F508'), ('on the same', r'on the same'),
    ('both mutations', r'both mutation'), ('double mutant', r'double mutant|double mutation'),
    ('accompanied by', r'accompanied by'), ('in combination', r'in combination|combined with'),
    ('co-occur', r'co-?occur'), ('haplotype', r'haplotype'), ('in association with', r'in association with'),
    ('also carries/found', r'also (?:carr|present|found|had|has|contain)'),
    ('carries also/both', r'carr(?:y|ies|ied|ying) (?:also|both)'), ('concomitant', r'concomitant'),
    ('plus', r'\bplus\b'), ('with <MUT>', r'with [A-Z]\d{2,4}[A-Z](?:\b|X)'),
    ('and <MUT>', r'and [A-Z]\d{2,4}[A-Z](?:\b|X)'), ('second mutation', r'second (?:mutation|change|variant)'),
    ('two mutations', r'two mutation'), ('linked to', r'linked (?:to|with)'),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--cftr2', default=os.environ.get('CFTR2_XLSX'), help='path to CFTR2 variant xlsx')
    ap.add_argument('--batch1', type=int, default=12, help='keyword-candidate batch size')
    ap.add_argument('--batch2', type=int, default=20, help='non-keyword batch size')
    args = ap.parse_args()

    os.makedirs(INTER, exist_ok=True)
    cftr1 = json.load(open(MUTATIONS))
    index, cftr2 = build_index(cftr1, args.cftr2)
    json.dump(index, open(os.path.join(INTER, 'alias_index.json'), 'w'))
    json.dump(cftr2, open(os.path.join(INTER, 'cftr2.json'), 'w'))

    rx = [(name, re.compile(p, re.I)) for name, p in KEYWORD_PATTERNS]
    allrecs, cands = [], []
    for m in cftr1:
        o = m.get('other_details') or ''
        ph = m.get('phenotype') or []
        p = ' || '.join(ph)
        hits = [name for name, r in rx if r.search(o) or r.search(p)]
        rec = {'sp_id': m['sp_id'], 'cdna_name': m.get('cdna_name'), 'protein_name': m.get('protein_name'),
               'legacy_name': m.get('legacy_name'), 'region': m.get('region'),
               'nucleotide_change': re.sub(r'^c\.', '', m.get('cdna_name') or ''),
               'consequence': derive_consequence(m), 'other_details': o, 'phenotype': ph}
        allrecs.append({**rec, 'keyword_hit': bool(hits)})
        if hits:
            cands.append({**rec, 'keywords': hits})

    json.dump(allrecs, open(os.path.join(INTER, 'all_records.json'), 'w'), ensure_ascii=False)
    json.dump(cands, open(os.path.join(INTER, 'candidates.json'), 'w'), ensure_ascii=False, indent=1)

    def has_text(r):
        return bool((r.get('other_details') or '').strip() or ' '.join(r.get('phenotype') or []).strip())

    cand_ids = {c['sp_id'] for c in cands}
    non = [r for r in allrecs if r['sp_id'] not in cand_ids]
    withtext = [r for r in non if has_text(r)]
    notext = [r for r in non if not has_text(r)]
    json.dump([r['sp_id'] for r in notext], open(os.path.join(INTER, 'notext_ids.json'), 'w'))

    # batches for the AI judgement workflows
    keys = ('sp_id', 'cdna_name', 'protein_name', 'legacy_name', 'region', 'nucleotide_change',
            'consequence', 'other_details', 'phenotype')
    for sub, data, size in [('batches', cands, args.batch1), ('batches2', withtext, args.batch2)]:
        d = os.path.join(INTER, sub); shutil.rmtree(d, ignore_errors=True); os.makedirs(d)
        n = 0
        for i in range(0, len(data), size):
            slim = [{k: r[k] for k in keys} for r in data[i:i + size]]
            prefix = 'batch_' if sub == 'batches' else 'b2_'
            json.dump(slim, open(os.path.join(d, f'{prefix}{n:02d}.json'), 'w'), ensure_ascii=False, indent=1)
            n += 1
        print(f'{sub}: {n} batches covering {len(data)} variants')

    print(f'total={len(cftr1)} candidates={len(cands)} non-keyword-with-text={len(withtext)} no-text={len(notext)}')
    print(f'alias index keys={len(index)}')


if __name__ == '__main__':
    main()
