"""
Step 2 of the CFTR1 complex-allele screen: turn the AI judgement passes into the
final table.

Inputs (../intermediate/): all_records.json, candidates.json, notext_ids.json,
  cftr2.json, judgements_pass1.json, judgements_pass2.json, judgements_pass3.json.
  (The judgements_*.json are the structured outputs of the three AI workflows in
  scripts/workflows/. all_records/candidates/notext_ids come from step 1.)

Outputs (../results/): CFTR1_complex_allele_screen.xlsx (2 sheets: hits + all),
  CFTR1_complex_alleles_HITS.csv, CFTR1_complex_alleles_ALL_2121.csv,
  CFTR1_complex_allele_screen_FULL.json, CFTR1_complex_allele_HITS.json.

Pipeline:
  pass1 (keyword candidates) + pass2 (non-keyword) give the initial verdict,
  keyword, referring text and reasoning; pass3 re-extracts CIS-only partners and
  reaffirms verdicts for the 177 hits; partner tokens are resolved to cDNA via
  lib_resolver; combos are assembled primary-first; reciprocal Yes pairs are
  made symmetric.
"""
import json, os, re, csv
import lib_resolver
from lib_resolver import resolve, norm_key

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
INTER = os.path.join(HERE, "..", "intermediate")
RESULTS = os.path.join(HERE, "..", "results")


def _derive_consequence(m):
    c = (m.get('cdna_name') or ''); p = (m.get('protein_name') or ''); region = (m.get('region') or '')
    if 'fs' in p or 'fsX' in p or 'fsTer' in p: return 'frameshift'
    if re.search(r'X\d*$', p) or 'Ter' in p or p.endswith('*'):
        return 'nonsense' if '>' in c else 'nonsense/truncating'
    if p.endswith('=') or 'p.(=)' in p or p == 'p.=': return 'synonymous'
    mm = re.match(r'p\.([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})', p)
    if mm and mm.group(1) != mm.group(3): return 'missense'
    if 'del' in c and 'ins' in c: return 'indel'
    if 'del' in c: return 'deletion (in-frame)' if ('del' in p and 'fs' not in p) else 'deletion'
    if 'ins' in c: return 'insertion'
    if 'dup' in c: return 'duplication'
    if re.search(r'c\.[-\d]+[-+]\d+', c) or region.startswith('intron'): return 'splicing (putative)'
    if '>' in c and not p: return 'substitution (non-coding/unknown)'
    if '>' in c: return 'substitution'
    return 'other/unspecified'


# Build the per-variant records straight from the CFTR1 archive data in this repo
# (regenerated here so step 2 runs standalone from the committed judgements).
_cftr1 = json.load(open(os.path.join(REPO, 'data', 'mutations.json')))
byid = {}
for _m in _cftr1:
    byid[_m['sp_id']] = {
        'sp_id': _m['sp_id'], 'cdna_name': _m.get('cdna_name'), 'protein_name': _m.get('protein_name'),
        'legacy_name': _m.get('legacy_name'), 'region': _m.get('region'),
        'nucleotide_change': re.sub(r'^c\.', '', _m.get('cdna_name') or ''),
        'consequence': _derive_consequence(_m),
        'other_details': _m.get('other_details') or '', 'phenotype': _m.get('phenotype') or []}

# CFTR2 clinical determination is an optional annotation (needs intermediate/cftr2.json from step 1).
_cftr2_path = os.path.join(INTER, 'cftr2.json')
cftr2 = json.load(open(_cftr2_path)) if os.path.exists(_cftr2_path) else []
det_by_cdna = {r['cdna']: r.get('determination') for r in cftr2 if r.get('cdna')}

COLS = ['Variant (cDNA name)', 'Protein Name', 'Legacy name', 'Nucleotide change', 'Consequence',
        'Key Word', 'Text referring to', 'Complex allele found or not', 'Your judgement reason',
        'Complex allele cDNA name', 'Trans partner (not part of complex allele)', 'Confidence',
        'Partner tokens (raw)', 'Screen source', 'CFTR2 determination (this variant)', 'sp_id', 'source_url']
VMAP = {'yes': 'Yes', 'uncertain': 'Uncertain', 'no': 'No'}


# ---- initial combo from raw partner tokens (pass1/2) ------------------------
def _sig(tok):
    t = tok.strip()
    m = re.match(r'^([ACGT]) to ([ACGT]) at position (\d+)$', t, re.I)
    if m: return f'{m.group(3)}{m.group(1).upper()}>{m.group(2).upper()}'
    s = re.sub(r'\s+', '', t.replace('->', '>').replace('/', '>')).upper()
    m = re.match(r'^C?\.?(\d+)([ACGT])>([ACGT])$', s)
    if m: return f'{m.group(1)}{m.group(2)}>{m.group(3)}'
    return norm_key(t)


def build_combo(prim, verdict, partner_tokens):
    resolved, unresolved_sig = [], {}
    for t in partner_tokens or []:
        t = (t or '').strip()
        if not t: continue
        r = resolve(t)
        if r and r != prim:
            if r not in resolved: resolved.append(r)
        elif not r:
            unresolved_sig.setdefault(_sig(t), t)
    unresolved = list(unresolved_sig.values())
    prim_is_complex = (';' in prim) or ('[' in prim and ']' in prim)
    if verdict in ('yes', 'uncertain'):
        if prim_is_complex and not resolved: return prim
        if resolved: return ';'.join([prim] + resolved)
        if unresolved: return ';'.join([prim] + [f'[{u}]' for u in unresolved])
        return prim if prim_is_complex else ''
    return ''


def assemble(judgements, source):
    rows = {}
    for j in judgements:
        sp = j['sp_id']; rec = byid.get(sp)
        if not rec: continue
        prim = rec['cdna_name']; verdict = j['is_complex_allele']
        rows[sp] = {
            'Variant (cDNA name)': prim, 'Protein Name': rec.get('protein_name'),
            'Legacy name': rec.get('legacy_name'), 'Nucleotide change': rec.get('nucleotide_change'),
            'Consequence': rec.get('consequence'), 'Key Word': j.get('keyword'),
            'Text referring to': j.get('text_referring_to'),
            'Complex allele found or not': VMAP.get(verdict, verdict),
            'Your judgement reason': j.get('reason'),
            'Complex allele cDNA name': build_combo(prim, verdict, j.get('partner_tokens')),
            'Trans partner (not part of complex allele)': '',
            'Confidence': j.get('confidence'),
            'Partner tokens (raw)': '; '.join([t for t in (j.get('partner_tokens') or []) if t.strip()]),
            'Screen source': source, 'CFTR2 determination (this variant)': det_by_cdna.get(prim, ''),
            'sp_id': sp, 'source_url': f'http://www.genet.sickkids.on.ca/MutationDetailPage.external?sp={sp}',
        }
    return rows


# ---- cis-only combo from pass3 ---------------------------------------------
def build_cis_combo(prim, cis_partners):
    prim_is_complex = ('[' in prim and ']' in prim) or (';' in prim)
    resolved, unresolved = [], []
    for cp in cis_partners:
        aw = cp.get('as_written', '').strip(); best = cp.get('best_hgvs_or_name', '').strip()
        r = None
        for cand in [best, aw, re.sub(r'\(.*?\)', '', best), re.sub(r'\(.*?\)', '', aw)]:
            if cand:
                r = resolve(cand)
                if r: break
        if r and r != prim:
            if r not in resolved: resolved.append(r)
        else:
            nm = best or aw
            if nm and nm not in unresolved and nm != prim: unresolved.append(nm)
    if prim_is_complex and not resolved and not unresolved: return prim
    return ';'.join([prim] + resolved + [f'[{u}]' for u in unresolved])


def cis_partner_cdnas(row):
    prim = row['Variant (cDNA name)']; combo = row['Complex allele cDNA name']
    if not combo or combo == prim: return []
    rest = combo[len(prim):].lstrip(';') if combo.startswith(prim) else combo
    return [p for p in rest.split(';') if p and not p.startswith('[') and p != prim]


def main():
    J1 = json.load(open(os.path.join(INTER, 'judgements_pass1.json')))
    J2 = json.load(open(os.path.join(INTER, 'judgements_pass2.json')))
    J3 = {r['sp_id']: r for r in json.load(open(os.path.join(INTER, 'judgements_pass3.json')))}
    judged = {j['sp_id'] for j in J1} | {j['sp_id'] for j in J2}
    notext_ids = [sp for sp, r in byid.items()
                  if sp not in judged and not (r['other_details'].strip() or ' '.join(r['phenotype']).strip())]

    merged = {}
    merged.update(assemble(J1, 'pass1: keyword screen'))
    merged.update(assemble(J2, 'pass2: AI (no keyword)'))
    for sp in notext_ids:
        rec = byid[sp]
        merged[sp] = {
            'Variant (cDNA name)': rec['cdna_name'], 'Protein Name': rec.get('protein_name'),
            'Legacy name': rec.get('legacy_name'), 'Nucleotide change': rec.get('nucleotide_change'),
            'Consequence': rec.get('consequence'), 'Key Word': 'none', 'Text referring to': '',
            'Complex allele found or not': 'No',
            'Your judgement reason': 'No Other Details or Submitted Phenotype text available; '
                                     'nothing to indicate a co-occurring variant.',
            'Complex allele cDNA name': '', 'Trans partner (not part of complex allele)': '',
            'Confidence': 'high', 'Partner tokens (raw)': '', 'Screen source': 'no-text (auto)',
            'CFTR2 determination (this variant)': '', 'sp_id': sp,
            'source_url': f'http://www.genet.sickkids.on.ca/MutationDetailPage.external?sp={sp}'}

    # apply pass3 cis-partner corrections to the hits
    for sp, corr in J3.items():
        if sp not in merged: continue
        row = merged[sp]; prim = row['Variant (cDNA name)']
        row['Complex allele found or not'] = VMAP.get(corr['verdict'], row['Complex allele found or not'])
        row['Your judgement reason'] = corr.get('reason', row['Your judgement reason'])
        row['Trans partner (not part of complex allele)'] = '; '.join(corr.get('trans_partners', []))
        if row['Complex allele found or not'] in ('Yes', 'Uncertain'):
            row['Complex allele cDNA name'] = build_cis_combo(prim, corr.get('cis_partners', []))
        else:
            row['Complex allele cDNA name'] = ''
        row['Screen source'] += ' + cis-corrected'

    # reciprocity: make Yes complex-allele pairs symmetric
    pmap = {}
    for row in merged.values():
        if row['Complex allele found or not'] == 'Yes':
            prim = row['Variant (cDNA name)']
            for b in cis_partner_cdnas(row):
                pmap.setdefault(prim, set()).add(b); pmap.setdefault(b, set()).add(prim)
    for row in merged.values():
        if row['Complex allele found or not'] == 'Yes':
            prim = row['Variant (cDNA name)']
            is_self = ('[' in prim and ']' in prim) or ';' in prim
            if not cis_partner_cdnas(row) and not is_self and prim in pmap:
                partners = sorted(p for p in pmap[prim] if p != prim)
                if partners:
                    row['Complex allele cDNA name'] = ';'.join([prim] + partners)

    rows = list(merged.values())
    order = {'Yes': 0, 'Uncertain': 1, 'No': 2}
    rows.sort(key=lambda r: (order.get(r['Complex allele found or not'], 3), r['sp_id']))
    hits = [r for r in rows if r['Complex allele found or not'] in ('Yes', 'Uncertain')]

    os.makedirs(RESULTS, exist_ok=True)
    _write_csv(rows, os.path.join(RESULTS, 'CFTR1_complex_alleles_ALL_2121.csv'))
    _write_csv(hits, os.path.join(RESULTS, 'CFTR1_complex_alleles_HITS.csv'))
    json.dump(rows, open(os.path.join(RESULTS, 'CFTR1_complex_allele_screen_FULL.json'), 'w'), ensure_ascii=False, indent=1)
    json.dump(hits, open(os.path.join(RESULTS, 'CFTR1_complex_allele_HITS.json'), 'w'), ensure_ascii=False, indent=1)
    _write_xlsx([('Complex allele hits', hits), ('All 2121 screened', rows)],
                os.path.join(RESULTS, 'CFTR1_complex_allele_screen.xlsx'))

    from collections import Counter
    print('verdicts:', dict(Counter(r['Complex allele found or not'] for r in rows)), '| hits:', len(hits))


def _write_csv(rows, path):
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=COLS, extrasaction='ignore')
        w.writeheader()
        for r in rows: w.writerow(r)


def _write_xlsx(sheets, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); first = True
    for name, rows in sheets:
        ws = wb.active if first else wb.create_sheet(); ws.title = name; first = False
        ws.append(COLS)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F4E78')
            c.alignment = Alignment(vertical='center', wrap_text=True)
        green = PatternFill('solid', fgColor='C6EFCE'); amber = PatternFill('solid', fgColor='FFEB9C')
        for r in rows:
            ws.append([r.get(c, '') for c in COLS])
            v = r.get('Complex allele found or not')
            if v == 'Yes':
                for c in ws[ws.max_row]: c.fill = green
            elif v == 'Uncertain':
                for c in ws[ws.max_row]: c.fill = amber
        widths = {'Text referring to': 55, 'Your judgement reason': 60, 'Complex allele cDNA name': 30,
                  'Trans partner (not part of complex allele)': 26, 'CFTR2 determination (this variant)': 26,
                  'source_url': 30, 'sp_id': 8, 'Confidence': 11}
        for i, c in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 18)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    wb.save(path)


if __name__ == '__main__':
    main()
